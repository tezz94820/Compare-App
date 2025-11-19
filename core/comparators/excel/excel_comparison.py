"""
Professional Excel Comparison Tool with Sheet-by-Sheet Analysis
Compares Excel files sheet by sheet with merged cell handling
"""

import openpyxl
import difflib
import os
import csv
import json
from html import escape
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Set
from collections import OrderedDict


class ExcelComparator:
    """Professional Excel comparison with sheet-by-sheet analytics."""
    
    def __init__(self, dev_excel: str, prod_excel: str, output_dir: str = "reports"):
        self.dev_excel = Path(dev_excel)
        self.prod_excel = Path(prod_excel)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.dev_sheets = OrderedDict()  # {sheet_name: sheet_data}
        self.prod_sheets = OrderedDict()
        self.sheet_diffs = []  # List of diffs for each sheet
        self.analytics = {}
        
    def extract_sheet_data(self, excel_path: Path) -> OrderedDict:
        """
        Extract data from all sheets, handling merged cells.
        Returns OrderedDict of {sheet_name: list of row strings}
        """
        sheets_data = OrderedDict()
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_rows = []
                
                # Get merged cell ranges
                merged_ranges = list(sheet.merged_cells.ranges)
                
                # Process each row
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    row_data = []
                    
                    for col_idx, cell in enumerate(row, start=1):
                        # Check if this cell is part of a merged range
                        cell_value = cell.value
                        
                        # If cell is merged, get the value from the top-left cell
                        for merged_range in merged_ranges:
                            if cell.coordinate in merged_range:
                                # Get the top-left cell of the merged range
                                top_left_cell = sheet.cell(
                                    merged_range.min_row, 
                                    merged_range.min_col
                                )
                                cell_value = top_left_cell.value
                                break
                        
                        # Convert cell value to string
                        if cell_value is None:
                            row_data.append("")
                        elif isinstance(cell_value, (int, float)):
                            row_data.append(str(cell_value))
                        else:
                            row_data.append(str(cell_value))
                    
                    # Join row data with tabs (to preserve column structure)
                    row_string = "\t".join(row_data).rstrip("\t")
                    
                    # Only add non-empty rows
                    if row_string.strip():
                        sheet_rows.append(row_string)
                
                sheets_data[sheet_name] = sheet_rows
            
            workbook.close()
            
        except Exception as e:
            print(f"❌ Error extracting data from {excel_path}: {e}")
            return OrderedDict()
        
        return sheets_data
    
    def compare_sheets(self):
        """Compare Excel files sheet by sheet."""
        
        # Get all unique sheet names
        all_sheet_names = set(self.dev_sheets.keys()) | set(self.prod_sheets.keys())
        
        for sheet_name in sorted(all_sheet_names):
            dev_data = self.dev_sheets.get(sheet_name, [])
            prod_data = self.prod_sheets.get(sheet_name, [])
            
            # Generate diff for this sheet
            differ = difflib.Differ()
            diff = list(differ.compare(dev_data, prod_data))
            
            self.sheet_diffs.append({
                'sheet_name': sheet_name,
                'exists_in_dev': sheet_name in self.dev_sheets,
                'exists_in_prod': sheet_name in self.prod_sheets,
                'dev_rows': len(dev_data),
                'prod_rows': len(prod_data),
                'diff': diff
            })
    
    def calculate_analytics(self) -> Dict:
        """Calculate comprehensive comparison analytics."""
        
        total_added = 0
        total_removed = 0
        total_changed = 0
        total_unchanged = 0
        
        for sheet_diff in self.sheet_diffs:
            diff = sheet_diff['diff']
            total_added += len([l for l in diff if l.startswith('+ ')])
            total_removed += len([l for l in diff if l.startswith('- ')])
            total_changed += len([l for l in diff if l.startswith('? ')]) // 2
            total_unchanged += len([l for l in diff if l.startswith('  ')])
        
        # Calculate overall similarity
        dev_text = "\n".join(["\n".join(data) for data in self.dev_sheets.values()])
        prod_text = "\n".join(["\n".join(data) for data in self.prod_sheets.values()])
        
        matcher = difflib.SequenceMatcher(None, dev_text, prod_text)
        similarity_ratio = matcher.ratio() if dev_text or prod_text else 1.0  # Raw ratio
        similarity = similarity_ratio * 100
        
        # Count cells (approximate by counting tabs + 1 per row)
        dev_cells = sum(row.count('\t') + 1 for sheet in self.dev_sheets.values() for row in sheet)
        prod_cells = sum(row.count('\t') + 1 for sheet in self.prod_sheets.values() for row in sheet)
        
        analytics = {
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'dev_file': self.dev_excel.name,
            'prod_file': self.prod_excel.name,
            'dev_size': self.dev_excel.stat().st_size if self.dev_excel.exists() else 0,
            'prod_size': self.prod_excel.stat().st_size if self.prod_excel.exists() else 0,
            'similarity_ratio': similarity_ratio,
            'similarity_percent': int(similarity),
            'difference_percent': int(100 - similarity),
            'total_sheets': {
                'dev': len(self.dev_sheets),
                'prod': len(self.prod_sheets),
                'max': max(len(self.dev_sheets), len(self.prod_sheets))
            },
            'changes': {
                'added': total_added,
                'removed': total_removed,
                'modified': total_changed,
                'unchanged': total_unchanged
            },
            'cells': {
                'dev': dev_cells,
                'prod': prod_cells,
                'diff': abs(dev_cells - prod_cells)
            }
        }
        
        return analytics
    
    def generate_html_report(self) -> str:
        """Generate professional HTML report with sheet-by-sheet comparison."""
        
        a = self.analytics
        
        # Generate sheet navigation buttons
        sheet_nav = ""
        for idx, sheet_diff in enumerate(self.sheet_diffs):
            sheet_name = sheet_diff['sheet_name']
            active_class = "active" if idx == 0 else ""
            sheet_nav += f'<button class="sheet-tab {active_class}" onclick="showSheet({idx})">{escape(sheet_name)}</button>'
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Comparison Report - {a['dev_file']} vs {a['prod_file']}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
            padding: 20px;
            color: #333;
            min-height: 100vh;
        }}
        
        .main-container {{
            max-width: 1800px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
            margin-bottom: 40px;
        }}
        
        .header {{
            background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 700;
        }}
        
        .header .subtitle {{
            font-size: 1.1em;
            opacity: 0.95;
        }}
        
        .analytics-dashboard {{
            padding: 40px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}
        
        .analytics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .metric-card {{
            background: white;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            border-left: 4px solid #2ecc71;
            transition: transform 0.2s;
        }}
        
        .metric-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }}
        
        .metric-label {{
            font-size: 0.9em;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }}
        
        .metric-value {{
            font-size: 2em;
            font-weight: 700;
            color: #2ecc71;
        }}
        
        .metric-subvalue {{
            font-size: 0.9em;
            color: #6c757d;
            margin-top: 5px;
        }}
        
        .similarity-bar {{
            width: 100%;
            height: 40px;
            background: #e9ecef;
            border-radius: 20px;
            overflow: hidden;
            position: relative;
            margin: 20px 0;
        }}
        
        .similarity-fill {{
            height: 100%;
            background: linear-gradient(90deg, #28a745, #20c997);
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 700;
            transition: width 1s ease;
        }}
        
        .legend {{
            display: flex;
            justify-content: center;
            gap: 30px;
            margin: 30px 0;
            flex-wrap: wrap;
        }}
        
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 10px 20px;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.08);
        }}
        
        .legend-color {{
            width: 20px;
            height: 20px;
            border-radius: 4px;
        }}
        
        .legend-added {{ background: #d4edda; }}
        .legend-removed {{ background: #f8d7da; }}
        .legend-changed {{ background: #fff3cd; }}
        
        .file-info {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin-top: 20px;
        }}
        
        .file-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        
        .file-card h3 {{
            color: #2ecc71;
            margin-bottom: 15px;
            font-size: 1.2em;
        }}
        
        .file-detail {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid #e9ecef;
        }}
        
        .file-detail:last-child {{
            border-bottom: none;
        }}
        
        .file-label {{
            color: #6c757d;
        }}
        
        .file-value {{
            font-weight: 600;
            color: #333;
        }}
        
        .sheet-navigation {{
            padding: 30px 40px 20px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
            overflow-x: auto;
            white-space: nowrap;
        }}
        
        .sheet-tabs {{
            display: inline-flex;
            gap: 10px;
        }}
        
        .sheet-tab {{
            padding: 12px 24px;
            background: white;
            border: 2px solid #e9ecef;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            transition: all 0.3s;
            color: #333;
        }}
        
        .sheet-tab:hover {{
            background: #e9ecef;
            transform: translateY(-2px);
        }}
        
        .sheet-tab.active {{
            background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
            color: white;
            border-color: #2ecc71;
        }}
        
        .sheets-container {{
            padding: 40px;
            background: #f8f9fa;
        }}
        
        .sheet-comparison {{
            display: none;
            background: #ffffff;
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            overflow: hidden;
            border: 2px solid #e9ecef;
        }}
        
        .sheet-comparison.active {{
            display: block;
        }}
        
        .sheet-header {{
            background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
            color: white;
            padding: 20px 30px;
            font-size: 1.3em;
            font-weight: 700;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .sheet-content {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 0;
        }}
        
        .sheet-column {{
            padding: 30px;
            background: #ffffff;
        }}
        
        .sheet-column:first-child {{
            border-right: 2px solid #e9ecef;
            background: #fafbfc;
        }}
        
        .sheet-column h3 {{
            color: #2ecc71;
            margin-bottom: 20px;
            font-size: 1.2em;
            padding-bottom: 10px;
            border-bottom: 2px solid #2ecc71;
            position: sticky;
            top: 0;
            background: inherit;
            z-index: 5;
        }}
        
        .content {{
            font-family: 'Courier New', 'Consolas', monospace;
            font-size: 13px;
            line-height: 1.6;
            white-space: pre-wrap;
            word-wrap: break-word;
        }}
        
        .line {{
            padding: 6px 10px;
            margin: 2px 0;
            border-radius: 4px;
            transition: all 0.15s;
        }}
        
        .line:hover {{
            opacity: 0.85;
        }}
        
        .added {{
            background-color: #d4edda;
            color: #155724;
            border-left: 4px solid #28a745;
            padding-left: 12px;
        }}
        
        .removed {{
            background-color: #f8d7da;
            color: #721c24;
            border-left: 4px solid #dc3545;
            padding-left: 12px;
        }}
        
        .changed {{
            background-color: #fff3cd;
            color: #856404;
            border-left: 4px solid #ffc107;
            padding-left: 12px;
        }}
        
        .empty-sheet {{
            color: #6c757d;
            font-style: italic;
            padding: 40px 20px;
            text-align: center;
            background: #f8f9fa;
            border-radius: 8px;
            border: 2px dashed #dee2e6;
        }}
        
        @media (max-width: 1200px) {{
            .sheet-content {{
                grid-template-columns: 1fr;
            }}
            
            .sheet-column:first-child {{
                border-right: none;
                border-bottom: 2px solid #e9ecef;
            }}
            
            .analytics-grid {{
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            }}
        }}
    </style>
</head>
<body>
    <div class="main-container">
        <div class="header">
            <h1>📊 Excel Comparison Report</h1>
            <div class="subtitle">{a['dev_file']} vs {a['prod_file']} • {a['timestamp']}</div>
        </div>
        
        <div class="analytics-dashboard">
            <h2 style="text-align: center; margin-bottom: 30px; color: #333; font-size: 1.8em;">
                📈 Analytics Dashboard
            </h2>
            
            <div class="analytics-grid">
                <div class="metric-card">
                    <div class="metric-label">Similarity Score</div>
                    <div class="metric-value">{a['similarity_percent']}%</div>
                    <div class="metric-subvalue">{a['difference_percent']}% different</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Total Sheets</div>
                    <div class="metric-value" style="color: #6c757d;">{a['total_sheets']['max']}</div>
                    <div class="metric-subvalue">Dev: {a['total_sheets']['dev']} | Prod: {a['total_sheets']['prod']}</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Rows Added</div>
                    <div class="metric-value" style="color: #28a745;">{a['changes']['added']}</div>
                    <div class="metric-subvalue">New content in Prod</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Rows Removed</div>
                    <div class="metric-value" style="color: #dc3545;">{a['changes']['removed']}</div>
                    <div class="metric-subvalue">Removed from Dev</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Rows Modified</div>
                    <div class="metric-value" style="color: #ffc107;">{a['changes']['modified']}</div>
                    <div class="metric-subvalue">Content changes</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Cells (Dev)</div>
                    <div class="metric-value" style="color: #17a2b8;">{a['cells']['dev']:,}</div>
                    <div class="metric-subvalue">Approximate count</div>
                </div>
                
                <div class="metric-card">
                    <div class="metric-label">Cells (Prod)</div>
                    <div class="metric-value" style="color: #17a2b8;">{a['cells']['prod']:,}</div>
                    <div class="metric-subvalue">Approximate count</div>
                </div>
            </div>
            
            <div class="similarity-bar">
                <div class="similarity-fill" style="width: {a['similarity_percent']}%">
                    {a['similarity_percent']}% Match
                </div>
            </div>
            
            <div class="file-info">
                <div class="file-card">
                    <h3>📄 Dev Excel</h3>
                    <div class="file-detail">
                        <span class="file-label">Filename:</span>
                        <span class="file-value">{a['dev_file']}</span>
                    </div>
                    <div class="file-detail">
                        <span class="file-label">File Size:</span>
                        <span class="file-value">{a['dev_size'] / 1024:.2f} KB</span>
                    </div>
                    <div class="file-detail">
                        <span class="file-label">Sheets:</span>
                        <span class="file-value">{a['total_sheets']['dev']}</span>
                    </div>
                </div>
                
                <div class="file-card">
                    <h3>📄 Prod Excel</h3>
                    <div class="file-detail">
                        <span class="file-label">Filename:</span>
                        <span class="file-value">{a['prod_file']}</span>
                    </div>
                    <div class="file-detail">
                        <span class="file-label">File Size:</span>
                        <span class="file-value">{a['prod_size'] / 1024:.2f} KB</span>
                    </div>
                    <div class="file-detail">
                        <span class="file-label">Sheets:</span>
                        <span class="file-value">{a['total_sheets']['prod']}</span>
                    </div>
                </div>
            </div>
            
            <div class="legend">
                <div class="legend-item">
                    <div class="legend-color legend-added"></div>
                    <span><strong>Added:</strong> {a['changes']['added']} rows</span>
                </div>
                <div class="legend-item">
                    <div class="legend-color legend-removed"></div>
                    <span><strong>Removed:</strong> {a['changes']['removed']} rows</span>
                </div>
                <div class="legend-item">
                    <div class="legend-color legend-changed"></div>
                    <span><strong>Modified:</strong> {a['changes']['modified']} rows</span>
                </div>
            </div>
        </div>
        
        <div class="sheet-navigation">
            <h3 style="margin-bottom: 15px; color: #333;">📑 Sheet Navigation</h3>
            <div class="sheet-tabs">
                {sheet_nav}
            </div>
        </div>
        
        <div class="sheets-container">"""
        
        # Generate comparison for each sheet
        for idx, sheet_data in enumerate(self.sheet_diffs):
            sheet_name = sheet_data['sheet_name']
            diff = sheet_data['diff']
            active_class = "active" if idx == 0 else ""
            
            html += f"""
            <div class="sheet-comparison {active_class}" id="sheet-{idx}">
                <div class="sheet-header">
                    📑 {escape(sheet_name)}
                </div>
                <div class="sheet-content">
                    <div class="sheet-column">
                        <h3>Dev Excel</h3>
                        <div class="content">"""
            
            # Check if sheet exists in dev
            if sheet_data['exists_in_dev'] and sheet_data['dev_rows'] > 0:
                # Generate Dev column content for this sheet
                for line in diff:
                    if line.startswith('- '):
                        html += f'<div class="line removed">{escape(line[2:])}</div>'
                    elif line.startswith('? '):
                        continue
                    elif line.startswith('+ '):
                        continue
                    else:
                        content = line[2:] if line.startswith('  ') else line
                        html += f'<div class="line">{escape(content)}</div>'
            else:
                html += '<div class="empty-sheet">📭 Sheet does not exist in Dev file</div>'
            
            html += """</div>
                    </div>
                    <div class="sheet-column">
                        <h3>Prod Excel</h3>
                        <div class="content">"""
            
            # Check if sheet exists in prod
            if sheet_data['exists_in_prod'] and sheet_data['prod_rows'] > 0:
                # Generate Prod column content for this sheet
                for line in diff:
                    if line.startswith('+ '):
                        html += f'<div class="line added">{escape(line[2:])}</div>'
                    elif line.startswith('? '):
                        continue
                    elif line.startswith('- '):
                        continue
                    else:
                        content = line[2:] if line.startswith('  ') else line
                        html += f'<div class="line">{escape(content)}</div>'
            else:
                html += '<div class="empty-sheet">📭 Sheet does not exist in Prod file</div>'
            
            html += """</div>
                    </div>
                </div>
            </div>"""
        
        html += """
        </div>
    </div>
    
    <script>
        function showSheet(sheetIndex) {
            // Hide all sheets
            const sheets = document.querySelectorAll('.sheet-comparison');
            sheets.forEach(sheet => sheet.classList.remove('active'));
            
            // Remove active class from all tabs
            const tabs = document.querySelectorAll('.sheet-tab');
            tabs.forEach(tab => tab.classList.remove('active'));
            
            // Show selected sheet
            document.getElementById('sheet-' + sheetIndex).classList.add('active');
            
            // Mark selected tab as active
            tabs[sheetIndex].classList.add('active');
        }
    </script>
</body>
</html>"""
        
        return html
    
    def compare(self) -> Tuple[str, Dict]:
        """Main comparison method - returns report path and analytics."""
        
        print(f"  🔍 Extracting data from Dev Excel...")
        self.dev_sheets = self.extract_sheet_data(self.dev_excel)
        
        print(f"  🔍 Extracting data from Prod Excel...")
        self.prod_sheets = self.extract_sheet_data(self.prod_excel)
        
        if not self.dev_sheets and not self.prod_sheets:
            print("  ❌ Error: Could not extract data from either Excel file")
            return "", {}
        
        print(f"  📑 Dev: {len(self.dev_sheets)} sheets | Prod: {len(self.prod_sheets)} sheets")
        
        print(f"  🔄 Comparing sheets...")
        self.compare_sheets()
        
        print(f"  📈 Calculating analytics...")
        self.analytics = self.calculate_analytics()
        
        print(f"  🎨 Generating HTML report...")
        html_report = self.generate_html_report()
        
        # Save report with sanitized filename
        safe_filename = f"{self.dev_excel.stem}_vs_{self.prod_excel.stem}".replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.output_dir / f"{safe_filename}_{timestamp}.html"
        
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_report)
        
        # Save analytics as JSON for summary
        analytics_path = self.output_dir / f"{safe_filename}_{timestamp}_analytics.json"
        with open(analytics_path, "w", encoding="utf-8") as f:
            json.dump(self.analytics, f, indent=2)
        
        print(f"  ✅ Report generated: {output_path.name}")
        
        return str(output_path.absolute()), self.analytics


class BatchExcelComparator:
    """Batch process Excel files using CSV file mapping."""
    
    def __init__(self, csv_file: str = "excel_mapping.csv", 
                 dev_folder: str = "dev", 
                 prod_folder: str = "prod", 
                 output_dir: str = "reports"):
        self.csv_file = Path(csv_file)
        self.dev_folder = Path(dev_folder)
        self.prod_folder = Path(prod_folder)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.file_mappings = []
        self.missing_files = []
        self.comparison_results = []
        
    def load_csv_mappings(self) -> bool:
        """Load file mappings from CSV file."""
        
        if not self.csv_file.exists():
            print(f"❌ CSV file not found: {self.csv_file}")
            print(f"   Please create a CSV file with columns: Sr.No, Dev Filename, Prod Filename")
            return False
        
        print(f"📋 Reading CSV file: {self.csv_file}")
        
        try:
            with open(self.csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                if reader.fieldnames is None:
                    print("❌ CSV file is empty")
                    return False
                
                # Normalize column names
                fieldnames = [col.strip().lower() for col in reader.fieldnames]
                
                dev_col = None
                prod_col = None
                
                for col in fieldnames:
                    if 'dev' in col and 'filename' in col:
                        dev_col = reader.fieldnames[fieldnames.index(col)]
                    elif 'prod' in col and 'filename' in col:
                        prod_col = reader.fieldnames[fieldnames.index(col)]
                
                if not dev_col or not prod_col:
                    print("❌ CSV must have 'Dev Filename' and 'Prod Filename' columns")
                    print(f"   Found columns: {', '.join(reader.fieldnames)}")
                    return False
                
                row_count = 0
                for row in reader:
                    dev_filename = row[dev_col].strip()
                    prod_filename = row[prod_col].strip()
                    
                    if dev_filename and prod_filename:
                        self.file_mappings.append({
                            'dev': dev_filename,
                            'prod': prod_filename
                        })
                        row_count += 1
                
                print(f"   ✅ Loaded {row_count} file mappings")
                return True
                
        except Exception as e:
            print(f"❌ Error reading CSV file: {e}")
            return False
    
    def validate_files(self):
        """Validate that all files in CSV exist in their respective folders."""
        
        print(f"\n🔍 Validating file existence...")
        
        valid_mappings = []
        
        for mapping in self.file_mappings:
            dev_path = self.dev_folder / mapping['dev']
            prod_path = self.prod_folder / mapping['prod']
            
            dev_exists = dev_path.exists()
            prod_exists = prod_path.exists()
            
            if dev_exists and prod_exists:
                valid_mappings.append({
                    'dev': mapping['dev'],
                    'prod': mapping['prod'],
                    'dev_path': dev_path,
                    'prod_path': prod_path
                })
            else:
                error_msg = []
                if not dev_exists:
                    error_msg.append(f"Dev file missing: {mapping['dev']}")
                if not prod_exists:
                    error_msg.append(f"Prod file missing: {mapping['prod']}")
                
                self.missing_files.append({
                    'dev': mapping['dev'],
                    'prod': mapping['prod'],
                    'error': ' | '.join(error_msg)
                })
        
        self.file_mappings = valid_mappings
        
        print(f"   ✅ Valid file pairs: {len(valid_mappings)}")
        if self.missing_files:
            print(f"   ⚠️  Missing/Invalid: {len(self.missing_files)}")
            for missing in self.missing_files:
                print(f"      ❌ {missing['error']}")
    
    def compare_all(self):
        """Compare all Excel pairs from CSV mapping."""
        
        if not self.load_csv_mappings():
            return
        
        if not self.file_mappings:
            print("\n❌ No file mappings found in CSV!")
            return
        
        self.validate_files()
        
        if not self.file_mappings:
            print("\n❌ No valid file pairs to compare!")
            return
        
        print(f"\n🔄 Starting batch comparison of {len(self.file_mappings)} Excel pairs...\n")
        
        for idx, mapping in enumerate(self.file_mappings, 1):
            print(f"[{idx}/{len(self.file_mappings)}] Comparing:")
            print(f"   Dev:  {mapping['dev']}")
            print(f"   Prod: {mapping['prod']}")
            
            comparator = ExcelComparator(
                str(mapping['dev_path']), 
                str(mapping['prod_path']), 
                str(self.output_dir)
            )
            
            report_path, analytics = comparator.compare()
            
            if report_path:
                self.comparison_results.append({
                    'dev_filename': mapping['dev'],
                    'prod_filename': mapping['prod'],
                    'report_path': report_path,
                    'analytics': analytics
                })
            
            print()
        
        # Generate summary
        print("=" * 80)
        print("📊 BATCH COMPARISON SUMMARY")
        print("=" * 80)
        print(f"\n✅ Successfully compared: {len(self.comparison_results)} Excel pairs")
        print(f"📁 Reports saved to: {self.output_dir.absolute()}\n")
        
        for result in self.comparison_results:
            a = result['analytics']
            print(f"📄 {result['dev_filename']} ↔ {result['prod_filename']}")
            print(f"   Similarity: {a['similarity_percent']}% | "
                  f"Added: {a['changes']['added']} | "
                  f"Removed: {a['changes']['removed']} | "
                  f"Modified: {a['changes']['modified']}")
        
        if self.missing_files:
            print(f"\n⚠️  Skipped {len(self.missing_files)} pairs due to missing files")
        
        print(f"\n🌐 Now generating master summary report...")
        
        # Auto-generate summary
        from excel_generate_summary import ExcelSummaryGenerator
        summary_gen = ExcelSummaryGenerator(str(self.output_dir))
        summary_path = summary_gen.generate_summary()
        
        if summary_path:
            print(f"✅ Master summary generated: {summary_path}")


def main():
    """Main entry point for batch comparison."""
    
    print("="*80)
    print("🚀 EXCEL BATCH COMPARISON TOOL (CSV-Based)")
    print("="*80)
    
    batch = BatchExcelComparator(
        csv_file="input/mappings/excel_file_mapping.csv",
        dev_folder="input/dev/excel",
        prod_folder="input/prod/excel",
        output_dir="reports/excel"
    )
    
    batch.compare_all()
    
    print("\n" + "="*80)
    print("✅ BATCH COMPARISON COMPLETE!")
    print("="*80)


if __name__ == "__main__":
    main()